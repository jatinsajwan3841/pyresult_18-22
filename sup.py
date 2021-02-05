from prettytable import PrettyTable

x = PrettyTable()

class result:
    def __init__(self, name, branch):
        self.name = name
        self.branch = branch
        self.xval = []
        self.yval = []
        self.total_marks = [0, 0]
        from openpyxl import Workbook
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.excelfiles = ['dat\\B. TECH. I SEM DEC 18.xlsx', 'dat\\B. TECH. II SEM JUNE 2019.xlsx',
                           'dat\\B. TECH. III SEM DECEMBER 2019.xlsx', 'dat\\B. TECH. IV SEM DECEMBER 2020.xlsx']
        self.letter = 'a'
        self.head = 1
        self.body = 4
        self.select()

    def select(self):
        from openpyxl import load_workbook
        for self.sem in range(0, 4):
            self.semn = load_workbook(
                self.excelfiles[self.sem], data_only=True)
            self.cS = self.semn[self.branch]

            self.search()

            if self.letter == 'a':
                print('data not matching')
                if self.sem == 2:
                    break
                continue

            self.vals()

            for row in range(4, 7):  # saving values to excel file
                for column in range(1, self.cS.max_column + 1):
                    self.sheet.cell(self.head, column).value = self.cS.cell(
                        row, column).value
                self.head += 1

            for column in range(1, self.cS.max_column + 1):
                self.sheet.cell(self.body, column).value = self.cS.cell(
                    self.letter, column).value

            self.head += 3
            self.body += 6
            self.wb.save("result.xlsx")

    def search(self):   # searching name
        for row in range(7, self.cS.max_row + 1):
            for column in "DE":
                self.cell_name = "{}{}".format(column, row)
                ex = self.cS[self.cell_name].value
                if type(ex) == str:
                    ex = ex.strip()
                    ex = ex.lower()
                if ex == self.name:
                    self.letter = row
                    return 0

    def vals(self):                          # saving values for table and plot
        row = 4
        for column in range(1, self.cS.max_column + 1):
            tem = self.cS.cell(row, column).value
            if type(tem) == str:
                tem = tem.strip()
                tem = tem.lower()
                if tem == 'result':
                    column -= 1
                    row += 2
                    self.sem += 1
                    x.field_names = ["Sem", "Marks", "Percentage"]
                    self.yval.append(self.cS.cell(self.letter, column).value / self.cS.cell(row,
                                                                                            column).value * 100)
                    self.xval.append(self.sem)
                    x.add_row([self.sem, "{}/{}".format(self.cS.cell(self.letter, column).value, self.cS.cell(row, column).value),
                                "{} %".format(round(self.cS.cell(self.letter, column).value / self.cS.cell(row, column).value * 100, 4))])
                    self.total_marks[0] = self.total_marks[0] + self.cS.cell(self.letter, column).value
                    self.total_marks[1] = self.total_marks[1] + self.cS.cell(row, column).value
                    return 0

    def display(self):
        self.clear_screen()
        print("Hello ", self.name, '\n')
        x.add_row(["Total :", "{}/{}".format(self.total_marks[0], self.total_marks[1]),
                   "{} %".format(round(self.total_marks[0]/self.total_marks[1]*100, 4))])
        print(x)
        import matplotlib.pyplot as plt
        plt.plot(self.xval, self.yval)
        plt.ylim(1, 100)
        plt.xlabel('Semester')
        plt.ylabel('Obtained Percentage')
        plt.title('Academic performance')
        plt.show()
        x.clear()
        self.xval.clear
        self.yval.clear

    def clear_screen(self):
        import os
        os.system("cls")
